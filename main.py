"""
FlowGo-Excel-Automator
Versión mínima (3 archivos)
"""

import sys
import os
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def print_help():
    print("Uso:")
    print("  python main.py --task demo --input <ruta_csv> --output <ruta_xlsx>")
    print("")
    print("Ejemplo:")
    print("  python main.py --task demo --input examples/datos.csv --output examples/salida.xlsx")


def read_args():
    # Lector básico de argumentos
    args = {"--task": None, "--input": None, "--output": None}
    i = 1
    while i < len(sys.argv):
        if sys.argv[i] in args:
            if i + 1 < len(sys.argv):
                args[sys.argv[i]] = sys.argv[i + 1]
                i += 2
            else:
                i += 1
        else:
            i += 1
    return args


def task_demo(csv_path, xlsx_path):
    # Verificar archivo CSV
    if not os.path.exists(csv_path):
        print("No se encontró el archivo CSV:", csv_path)
        return

    # Leer CSV
    filas = []
    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        lector = csv.reader(f)
        encabezado = next(lector, None)
        for fila in lector:
            filas.append(fila)

    # Crear XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Encabezado
    if encabezado:
        for c, val in enumerate(encabezado, start=1):
            ws.cell(row=1, column=c, value=val)

    # Escribir filas
    start_row = 2 if encabezado else 1
    for r, fila in enumerate(filas, start=start_row):
        for c, val in enumerate(fila, start=1):
            if c == 2:
                try:
                    val = float(val)
                except:
                    pass
            ws.cell(row=r, column=c, value=val)

    # Calcular TOTAL de la segunda columna
    total = 0.0
    for r in range(start_row, start_row + len(filas)):
        try:
            total += float(ws.cell(row=r, column=2).value)
        except:
            pass

    total_row = start_row + len(filas)
    ws.cell(row=total_row + 1, column=1, value="TOTAL")
    ws.cell(row=total_row + 1, column=2, value=total)

    # Ajustar ancho de columnas
    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Guardar XLSX
    carpeta = os.path.dirname(xlsx_path)
    if carpeta:
        os.makedirs(carpeta, exist_ok=True)

    wb.save(xlsx_path)
    print("Archivo XLSX generado:", xlsx_path)


def main():
    args = read_args()
    if args["--task"] == "demo":
        if args["--input"] and args["--output"]:
            task_demo(args["--input"], args["--output"])
        else:
            print("Faltan argumentos --input y/o --output.")
            print_help()
    else:
        print("Tarea no válida o no especificada.")
        print_help()


if __name__ == "__main__":
    main()
